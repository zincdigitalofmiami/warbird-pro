#!/usr/bin/env python3
"""
Backfill GPR (Geopolitical Risk) data from Caldara-Iacoviello XLS.
Writes to geopolitical_risk_1d(ts, gpr_daily, gpr_threats, gpr_acts).
"""
from __future__ import annotations

import datetime
import os
import sys
from pathlib import Path

from supabase import create_client

def main() -> None:
    xls_path = Path(sys.argv[1] if len(sys.argv) > 1 else "data/gpr_web.xls")
    if not xls_path.exists():
        print(f"ERROR: {xls_path} not found. Download from matteoiacoviello.com/gpr.htm")
        sys.exit(1)

    sb_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ["SUPABASE_URL"]
    sb_key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    supabase = create_client(sb_url, sb_key)

    try:
        import xlrd
        book = xlrd.open_workbook(str(xls_path))
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
        rows_raw = [
            {headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)}
            for r in range(1, sheet.nrows)
        ]
    except ImportError:
        import openpyxl
        wb = openpyxl.load_workbook(str(xls_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows_raw = [
            {headers[i]: row[i].value for i in range(len(headers))}
            for row in ws.iter_rows(min_row=2)
        ]

    # Find columns: DAY (YYYYMMDD string), GPRD, GPRD_ACT, GPRD_THREAT
    day_col = next((h for h in headers if h and str(h).upper() == "DAY"), None)
    date_col = next((h for h in headers if h and str(h).lower() == "date"), None)
    gprd_col = next((h for h in headers if h and str(h).upper() == "GPRD"), None)
    acts_col = next((h for h in headers if h and "ACT" in str(h).upper()), None)
    threat_col = next((h for h in headers if h and "THREAT" in str(h).upper()), None)

    if not gprd_col:
        print(f"ERROR: can't find GPRD column. Headers: {headers}")
        sys.exit(1)

    rows = []
    for raw in rows_raw:
        # Parse date from DAY column (YYYYMMDD) or date column (Excel serial)
        ts = None
        if day_col and raw.get(day_col):
            day_str = str(raw[day_col]).strip()
            if len(day_str) == 8 and day_str.isdigit():
                ts = f"{day_str[:4]}-{day_str[4:6]}-{day_str[6:8]}T00:00:00Z"
        if not ts and date_col and raw.get(date_col):
            date_val = raw[date_col]
            if isinstance(date_val, (int, float)):
                d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(date_val))
                ts = d.isoformat() + "T00:00:00Z"
            else:
                ts = str(date_val)[:10] + "T00:00:00Z"

        if not ts:
            continue

        gprd = raw.get(gprd_col)
        if gprd is None:
            continue

        row = {
            "ts": ts,
            "gpr_daily": float(gprd),
        }
        if acts_col and raw.get(acts_col) is not None:
            row["gpr_acts"] = float(raw[acts_col])
        if threat_col and raw.get(threat_col) is not None:
            row["gpr_threats"] = float(raw[threat_col])

        rows.append(row)

    print(f"Upserting {len(rows)} GPR rows...")

    for i in range(0, len(rows), 500):
        chunk = rows[i:i + 500]
        supabase.table("geopolitical_risk_1d").upsert(chunk, on_conflict="ts").execute()

    print(f"Done. {len(rows)} rows upserted.")

if __name__ == "__main__":
    main()
